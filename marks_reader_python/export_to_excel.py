"""
Export class marks to Excel spreadsheet format
"""

import sys
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_to_excel(class_code, csv_dir=r'S:\Chn\classes\csv_exports_python', output_dir=r'S:\Chn\classes'):
    """Export class marks to Excel with formatting"""

    csv_file = f"{csv_dir}\\{class_code}_marks.csv"
    excel_file = f"{output_dir}\\{class_code}_marks.xlsx"

    # Read CSV data
    with open(csv_file, 'r') as f:
        reader = csv.reader(f)
        rows = list(reader)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = class_code

    # Write all data
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Header formatting
            if row_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Left align name, ID, homeform
                if col_idx <= 3:
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')

    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # Name
    ws.column_dimensions['B'].width = 12  # Student Number
    ws.column_dimensions['C'].width = 8   # Homeform

    # All other columns (use get_column_letter for columns beyond Z)
    from openpyxl.utils import get_column_letter
    for col in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9

    # Freeze first row and first 3 columns
    ws.freeze_panes = 'D2'

    # Save
    wb.save(excel_file)
    print(f"Saved to: {excel_file}")
    print(f"  {len(rows)-1} students")
    print(f"  {ws.max_column - 3} assignments + {ws.max_column - len(rows[0])} summary columns")


if __name__ == '__main__':
    class_code = sys.argv[1] if len(sys.argv) > 1 else 'TIK2O1-1'
    export_to_excel(class_code)
